"""
Flask 웹 애플리케이션 v6 - 다중 광고주 탭 UI
- 광고주별 탭 전환 (클라이언트 사이드)
- 각 광고주마다 독립된 상품/키워드/순위 관리
- 모든 버튼 AJAX 동작
- 매일 오전 11시 자동 순위 추적
"""
from flask import Flask, render_template, request, jsonify, redirect, url_for, flash
import threading, logging, os, re
from datetime import datetime
from db import init_db, get_conn
from engine import parse_product_info, track_client, search_shopping

from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
import pytz

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "naver_rank_agency_2025")

tracking_status = {}
global_tracking = {"running": False, "last_run": None}

# ────────────────────────────────────────────
# 공통 헬퍼
# ────────────────────────────────────────────
def get_api_keys():
    env_id     = os.environ.get("NAVER_CLIENT_ID", "")
    env_secret = os.environ.get("NAVER_CLIENT_SECRET", "")
    if env_id and env_secret:
        return env_id, env_secret
    conn = get_conn()
    rows = {r["key"]: r["value"] for r in
            conn.execute("SELECT key,value FROM settings WHERE key IN ('client_id','client_secret')").fetchall()}
    conn.close()
    return rows.get("client_id", ""), rows.get("client_secret", "")


def get_client_data(cid):
    """광고주 한 명의 rows(상품×키워드), products, keywords 반환"""
    conn = get_conn()
    c    = conn.cursor()

    # 상품 × 키워드 조합 + 최신 순위
    c.execute("""
        SELECT
            p.id          AS pid,
            p.product_id  AS product_id,
            p.product_name AS product_name,
            p.product_url  AS product_url,
            k.id           AS kid,
            k.keyword      AS keyword,
            rh.rank        AS rank,
            rh.lprice      AS lprice,
            rh.checked_at  AS checked_at
        FROM products p
        CROSS JOIN keywords k ON k.client_id = p.client_id
        LEFT JOIN rank_history rh ON (
            rh.product_id = p.product_id
            AND rh.client_id  = p.client_id
            AND rh.keyword    = k.keyword
            AND rh.id = (
                SELECT MAX(id) FROM rank_history
                WHERE product_id = p.product_id
                  AND client_id  = p.client_id
                  AND keyword    = k.keyword
            )
        )
        WHERE p.client_id=?
        ORDER BY p.id, k.id
    """, (cid,))
    combo_rows = [dict(r) for r in c.fetchall()]

    # 키워드 없는 상품 단독
    c.execute("""
        SELECT p.id AS pid, p.product_id, p.product_name, p.product_url
        FROM products p
        WHERE p.client_id=? AND NOT EXISTS (
            SELECT 1 FROM keywords k WHERE k.client_id=?
        )
    """, (cid, cid))
    for sp in c.fetchall():
        combo_rows.append({
            "pid": sp["pid"], "product_id": sp["product_id"],
            "product_name": sp["product_name"], "product_url": sp["product_url"],
            "kid": None, "keyword": "—",
            "rank": None, "lprice": None, "checked_at": None,
        })

    c.execute("SELECT id, keyword FROM keywords WHERE client_id=? ORDER BY id", (cid,))
    keywords = [dict(r) for r in c.fetchall()]

    c.execute("SELECT id, product_id, product_name, product_url FROM products WHERE client_id=? ORDER BY id", (cid,))
    products = [dict(r) for r in c.fetchall()]

    conn.close()
    return combo_rows, products, keywords


# ────────────────────────────────────────────
# 전체 추적
# ────────────────────────────────────────────
def run_all_tracking(source="manual"):
    if global_tracking["running"]:
        return
    api_id, api_secret = get_api_keys()
    if not api_id:
        logger.warning("[추적] API 키 미설정")
        return

    conn = get_conn()
    clients = conn.execute("SELECT id,name FROM clients").fetchall()
    conn.close()
    if not clients:
        return

    global_tracking["running"] = True
    logger.info(f"[추적 시작] {source} | {len(clients)}개 광고주")
    try:
        for cl in clients:
            cid = cl["id"]
            conn2 = get_conn()
            prods = [dict(r) for r in conn2.execute(
                "SELECT product_id,catalog_id,url_product_id,mall_name,product_name FROM products WHERE client_id=?", (cid,)
            ).fetchall()]
            kws = [r["keyword"] for r in conn2.execute(
                "SELECT keyword FROM keywords WHERE client_id=?", (cid,)
            ).fetchall()]
            conn2.close()
            if not prods or not kws:
                continue
            tracking_status[cid] = "running"
            try:
                results = track_client(api_id, api_secret, cid, prods, kws, max_pages=10)
                conn3 = get_conn()
                for r in results:
                    conn3.execute("""
                        INSERT INTO rank_history
                        (client_id,product_id,product_name,keyword,rank,
                         lprice,mall_name,product_type,matched_id,checked_at)
                        VALUES (?,?,?,?,?,?,?,?,?,?)
                    """, (r["client_id"], r["product_id"], r["product_name"],
                          r["keyword"], r["rank"], r.get("lprice"), r.get("mall_name"),
                          r.get("product_type"), r.get("matched_id"), r["checked_at"]))
                conn3.commit()
                conn3.close()
                tracking_status[cid] = "done"
                logger.info(f"  ✅ {cl['name']} 완료 ({len(results)}건)")
            except Exception as e:
                tracking_status[cid] = f"error:{e}"
                logger.error(f"  ❌ {cl['name']} 오류: {e}")
        global_tracking["last_run"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    finally:
        global_tracking["running"] = False


def scheduled_job():
    logger.info("⏰ 자동 스케줄 실행")
    threading.Thread(target=lambda: run_all_tracking("schedule"), daemon=True).start()


# ────────────────────────────────────────────
# APScheduler
# ────────────────────────────────────────────
KST = pytz.timezone("Asia/Seoul")
scheduler = BackgroundScheduler(timezone=KST)
scheduler.add_job(scheduled_job, CronTrigger(hour=11, minute=0, timezone=KST),
                  id="daily_track", replace_existing=True)
scheduler.start()
logger.info("⏰ 스케줄러 시작 — 매일 KST 11:00")


# ════════════════════════════════════════════
# 메인 대시보드
# ════════════════════════════════════════════
@app.route("/")
def index():
    conn = get_conn()
    clients = [dict(r) for r in conn.execute("SELECT id,name,memo FROM clients ORDER BY id").fetchall()]
    conn.close()

    # 광고주가 하나도 없으면 빈 상태로 시작 (자동 생성 X)
    client_data = {}
    for cl in clients:
        rows, products, keywords = get_client_data(cl["id"])
        client_data[cl["id"]] = {
            "rows": rows,
            "products": products,
            "keywords": keywords,
        }

    job = scheduler.get_job("daily_track")
    next_run = job.next_run_time.astimezone(KST).strftime("%m/%d %H:%M") if job and job.next_run_time else "-"

    return render_template("index.html",
                           clients=clients,
                           client_data=client_data,
                           global_tracking=global_tracking,
                           next_run=next_run)


# ════════════════════════════════════════════
# 광고주 CRUD (AJAX)
# ════════════════════════════════════════════
@app.route("/clients/add", methods=["POST"])
def add_client():
    name = (request.form.get("name") or request.json.get("name", "") if request.is_json else request.form.get("name","")).strip()
    if not name:
        return jsonify({"error": "광고주명을 입력하세요."}), 400
    conn = get_conn()
    try:
        conn.execute("INSERT INTO clients (name, memo) VALUES (?,?)", (name, ""))
        conn.commit()
        cid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500
    conn.close()
    return jsonify({"ok": True, "id": cid, "name": name})


@app.route("/clients/<int:cid>/delete", methods=["POST"])
def delete_client(cid):
    conn = get_conn()
    conn.execute("DELETE FROM rank_history WHERE client_id=?", (cid,))
    conn.execute("DELETE FROM keywords WHERE client_id=?", (cid,))
    conn.execute("DELETE FROM products WHERE client_id=?", (cid,))
    conn.execute("DELETE FROM clients WHERE id=?", (cid,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


# ════════════════════════════════════════════
# 상품 CRUD (AJAX)
# ════════════════════════════════════════════
@app.route("/clients/<int:cid>/products/add", methods=["POST"])
def add_product(cid):
    product_url  = request.form.get("product_url", "").strip()
    product_name = request.form.get("product_name", "").strip()
    if not product_url:
        return jsonify({"error": "URL을 입력하세요."}), 400

    # 클라이언트 존재 확인
    conn = get_conn()
    if not conn.execute("SELECT id FROM clients WHERE id=?", (cid,)).fetchone():
        conn.close()
        return jsonify({"error": "존재하지 않는 광고주입니다."}), 404

    info           = parse_product_info(product_url)
    product_id     = info.get("product_id", "") or product_url
    catalog_id     = info.get("catalog_id", "")
    url_product_id = info.get("product_id", "")
    display_name   = product_name or product_id

    try:
        conn.execute("""
            INSERT OR IGNORE INTO products
            (client_id,product_url,product_id,catalog_id,url_product_id,mall_name,product_name)
            VALUES (?,?,?,?,?,?,?)
        """, (cid, product_url, product_id, catalog_id, url_product_id, "", display_name))
        conn.commit()
        pid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500
    conn.close()
    return jsonify({"ok": True, "pid": pid, "product_id": product_id, "product_name": display_name, "product_url": product_url})


@app.route("/clients/<int:cid>/products/<int:pid>/delete", methods=["POST"])
def delete_product(cid, pid):
    conn = get_conn()
    conn.execute("DELETE FROM products WHERE id=? AND client_id=?", (pid, cid))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


# ════════════════════════════════════════════
# 키워드 CRUD (AJAX)
# ════════════════════════════════════════════
@app.route("/clients/<int:cid>/keywords/add", methods=["POST"])
def add_keyword(cid):
    raw = request.form.get("keyword", "").strip()
    if not raw:
        return jsonify({"error": "키워드를 입력하세요."}), 400

    conn = get_conn()
    if not conn.execute("SELECT id FROM clients WHERE id=?", (cid,)).fetchone():
        conn.close()
        return jsonify({"error": "존재하지 않는 광고주입니다."}), 404

    kws = [k.strip() for k in re.split(r"[,\n]+", raw) if k.strip()]
    added = []
    for kw in kws:
        try:
            conn.execute("INSERT OR IGNORE INTO keywords (client_id,keyword) VALUES (?,?)", (cid, kw))
            added.append(kw)
        except Exception:
            pass
    conn.commit()

    # 추가된 keyword id 목록
    kid_map = {}
    for kw in added:
        row = conn.execute("SELECT id FROM keywords WHERE client_id=? AND keyword=?", (cid, kw)).fetchone()
        if row:
            kid_map[kw] = row["id"]
    conn.close()
    return jsonify({"ok": True, "added": added, "kid_map": kid_map})


@app.route("/clients/<int:cid>/keywords/<int:kid>/delete", methods=["POST"])
def delete_keyword(cid, kid):
    conn = get_conn()
    conn.execute("DELETE FROM keywords WHERE id=? AND client_id=?", (kid, cid))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


# ════════════════════════════════════════════
# 추적
# ════════════════════════════════════════════
@app.route("/track/now", methods=["POST"])
def track_now():
    if global_tracking["running"]:
        return jsonify({"error": "이미 추적 중입니다."}), 409
    threading.Thread(target=lambda: run_all_tracking("manual"), daemon=True).start()
    return jsonify({"ok": True, "message": "추적을 시작했습니다."})


@app.route("/track/status")
def track_status():
    return jsonify({"running": global_tracking["running"], "last_run": global_tracking["last_run"]})


# ════════════════════════════════════════════
# 상품 검색 API
# ════════════════════════════════════════════
@app.route("/api/search-products")
def api_search_products():
    q    = request.args.get("q", "").strip()
    page = int(request.args.get("page", 1))
    if not q:
        return jsonify({"error": "검색어를 입력하세요."})
    api_id, api_secret = get_api_keys()
    if not api_id:
        return jsonify({"error": "Naver API 키를 먼저 설정해주세요."})
    try:
        start    = (page - 1) * 20 + 1
        data     = search_shopping(api_id, api_secret, q, display=20, start=start)
        items_raw = data.get("items", [])
        total     = data.get("total", 0)
        items = []
        for i, item in enumerate(items_raw):
            pid    = item.get("productId", "")
            ptype  = str(item.get("productType", "2"))
            is_cat = ptype == "1"
            link   = item.get("link", "")
            items.append({
                "rank":      start + i,
                "productId": pid,
                "title":     re.sub(r"<[^>]+>", "", item.get("title", "")),
                "mallName":  item.get("mallName", ""),
                "lprice":    int(item.get("lprice", 0) or 0),
                "image":     item.get("image", ""),
                "isCatalog": is_cat,
                "link":      link,
                "addUrl":    f"https://search.shopping.naver.com/catalog/{pid}" if is_cat else link,
            })
        return jsonify({"items": items, "total": total})
    except Exception as e:
        return jsonify({"error": str(e)})


# ════════════════════════════════════════════
# 히스토리 API
# ════════════════════════════════════════════
@app.route("/api/history")
def api_history():
    product_id = request.args.get("pid", "")
    keyword    = request.args.get("kw", "")
    conn = get_conn()
    rows = [{"rank": r["rank"], "date": r["checked_at"][:16]} for r in conn.execute("""
        SELECT rank, checked_at FROM rank_history
        WHERE product_id=? AND keyword=?
          AND checked_at >= datetime('now','-30 days','localtime')
        ORDER BY checked_at ASC LIMIT 60
    """, (product_id, keyword)).fetchall()]
    conn.close()
    return jsonify(rows)


# ════════════════════════════════════════════
# 설정
# ════════════════════════════════════════════
@app.route("/settings", methods=["GET", "POST"])
def settings():
    if request.method == "POST":
        cid    = request.form.get("client_id", "").strip()
        secret = request.form.get("client_secret", "").strip()
        conn   = get_conn()
        conn.execute("INSERT OR REPLACE INTO settings (key,value) VALUES ('client_id',?)", (cid,))
        conn.execute("INSERT OR REPLACE INTO settings (key,value) VALUES ('client_secret',?)", (secret,))
        conn.commit()
        conn.close()
        flash("API 키가 저장되었습니다.", "success")
        return redirect(url_for("settings"))
    api_id, api_secret = get_api_keys()
    return render_template("settings.html", client_id=api_id, client_secret=api_secret)



# ════════════════════════════════════════════
# 업무 자동화 페이지
# ════════════════════════════════════════════
@app.route("/automation")
def automation():
    return render_template("automation.html")


# ════════════════════════════════════════════
# 업무 자동화 API 섹션 (v15 구조 대응)
# 컬럼: A(구분), B(WEB), C(시작일), D(종료일),
#       E(총예산), F(총예산수식), G(일예산수식), H(포인트10수식),
#       I(검색어=PID앞3), J(미션내용), K(정답=PID앞5),
#       L(힌트URL), M(업체명), N(일유입목표), O(글자수수식)
# 병합: A,C,D,E,F,G,H,N,O / 수식보존: E,F,G,H,O
# 데이터 시작행: 3 (헤더 2행)
# ════════════════════════════════════════════

@app.route("/api/fetch-store-name")
def api_fetch_store_name():
    """업체명 조회: keyword+PID로 Shopping API mallName 우선, 실패시 크롤링
    ?slug=xxx&pid=123&keyword=장뇌삼
    """
    import requests as req
    slug    = request.args.get("slug", "").strip()
    pid     = request.args.get("pid", "").strip()
    keyword = request.args.get("keyword", "").strip()

    if not slug:
        return jsonify({"ok": False, "name": "", "error": "slug 없음"})

    client_id, client_secret = get_api_keys()
    headers_naver = {
        "X-Naver-Client-Id":     client_id,
        "X-Naver-Client-Secret": client_secret,
    }

    def shopping_mall_name(query, pid_to_match):
        """Shopping API로 query 검색, pid_to_match 매칭 후 mallName 반환"""
        if not client_id or not query:
            return None
        try:
            import urllib.parse
            api_url = (f"https://openapi.naver.com/v1/search/shop.json"
                       f"?query={urllib.parse.quote(query)}&display=30&sort=sim")
            resp = req.get(api_url, headers=headers_naver, timeout=8)
            if resp.status_code != 200:
                return None
            items = resp.json().get("items", [])
            for item in items:
                item_pid = str(item.get("productId", ""))
                link = item.get("link", "")
                m = re.search(r'/products/(\d+)', link)
                link_pid = m.group(1) if m else ""
                if pid_to_match and (pid_to_match == item_pid or pid_to_match == link_pid):
                    mall = item.get("mallName", "").strip()
                    if mall and len(mall) <= 30:
                        return mall
        except Exception as e:
            logger.warning(f"[ShoppingAPI] query={query} 오류: {e}")
        return None

    # 1) keyword + PID 매칭 → mallName (가장 정확)
    if keyword and pid:
        name = shopping_mall_name(keyword, pid)
        if name:
            logger.info(f"[StoreNameFetch] kw-match slug={slug} kw={keyword} -> {name}")
            return jsonify({"ok": True, "name": name, "slug": slug, "source": "api-kw"})

    # 2) slug로 검색 + PID 매칭 → mallName
    if slug and pid:
        name = shopping_mall_name(slug, pid)
        if name:
            logger.info(f"[StoreNameFetch] slug-match slug={slug} -> {name}")
            return jsonify({"ok": True, "name": name, "slug": slug, "source": "api-slug"})

    # 3) 크롤링 폴백
    crawl_headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                      "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept-Language": "ko-KR,ko;q=0.9",
        "Accept": "text/html,application/xhtml+xml",
    }
    try:
        profile_url = f"https://smartstore.naver.com/{slug}"
        resp = req.get(profile_url, headers=crawl_headers, timeout=8, allow_redirects=True)
        html = resp.text
        for pat in [
            r'<meta[^>]+property=["\']og:title["\'][^>]+content=["\']([^"\']+)["\']',
            r'<meta[^>]+content=["\']([^"\']+)["\'][^>]+property=["\']og:title["\']',
        ]:
            m = re.search(pat, html)
            if m:
                raw = m.group(1).strip()
                name = raw.split(':')[0].strip() if ':' in raw else raw
                if name and len(name) <= 30:
                    logger.info(f"[StoreNameFetch] crawl-og slug={slug} -> {name}")
                    return jsonify({"ok": True, "name": name, "slug": slug, "source": "crawl"})
        m_title = re.search(r'<title[^>]*>([^<]+)</title>', html)
        if m_title:
            raw = m_title.group(1).strip()
            name = raw.split(':')[0].strip() if ':' in raw else raw
            if name and len(name) <= 30:
                logger.info(f"[StoreNameFetch] crawl-title slug={slug} -> {name}")
                return jsonify({"ok": True, "name": name, "slug": slug, "source": "crawl-title"})
    except Exception as e:
        logger.warning(f"[StoreNameFetch] crawl 실패: {e}")

    # 4) 최종 폴백: slug 반환
    logger.info(f"[StoreNameFetch] fallback slug={slug}")
    return jsonify({"ok": False, "name": slug, "slug": slug, "source": "slug"})


@app.route("/api/check-rank", methods=["POST"])
def api_check_rank():
    """키워드별 네이버 쇼핑 순위 확인 — top15 이내 여부 반환"""
    import requests as req
    data  = request.get_json(force=True)
    jobs  = data.get("jobs", [])   # [{pid, url, keywords:[]}]
    client_id, client_secret = get_api_keys()

    if not client_id:
        return jsonify({"ok": False, "error": "API 키가 설정되지 않았습니다."})

    def get_rank(pid, keyword):
        """Naver Shopping API로 PID 순위 조회"""
        headers = {
            "X-Naver-Client-Id":     client_id,
            "X-Naver-Client-Secret": client_secret,
        }
        try:
            url = (f"https://openapi.naver.com/v1/search/shop.json"
                   f"?query={req.utils.quote(keyword)}&display=30&sort=sim")
            resp = req.get(url, headers=headers, timeout=8)
            if resp.status_code != 200:
                return None
            items = resp.json().get("items", [])
            for i, item in enumerate(items):
                # productId 또는 link에서 PID 추출
                item_pid = str(item.get("productId", ""))
                link = item.get("link", "")
                m = re.search(r'/products/(\d+)', link)
                link_pid = m.group(1) if m else ""
                if pid and (pid == item_pid or pid == link_pid):
                    return i + 1
            return 0  # 30위 이내 없음
        except Exception as e:
            logger.warning(f"[CheckRank] kw={keyword} 오류: {e}")
            return None

    results = []
    for job in jobs:
        pid = str(job.get("pid", ""))
        kws = job.get("keywords", [])
        url = job.get("url", "")
        kw_results = []
        for kw in kws:
            rank = get_rank(pid, kw) if kw else None
            kw_results.append({"keyword": kw, "rank": rank})
        results.append({"pid": pid, "url": url, "keywords": kw_results})

    return jsonify({"ok": True, "results": results})


# ════════════════════════════════════════════
# Excel 새 파일 내보내기 (v15 구조)
# ════════════════════════════════════════════
@app.route("/api/automation/excel-export", methods=["POST"])
def api_excel_export():
    """캠페인 데이터를 새 .xlsx로 내보내기 — v15 컬럼 구조
    A(1)=구분, B(2)=WEB, C(3)=시작일, D(4)=종료일,
    E(5)=총예산, F(6)=총예산수식, G(7)=일예산수식, H(8)=포인트(수식),
    I(9)=검색어, J(10)=미션내용, K(11)=정답,
    L(12)=힌트URL, M(13)=업체명, N(14)=일유입목표, O(15)=글자수수식
    병합: A,C,D,E,F,G,H,N,O → 5행씩
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    data = request.get_json(force=True)
    campaigns = data.get("campaigns", [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "네이버쇼핑캠페인"

    # 헤더 스타일
    hdr_fill  = PatternFill("solid", fgColor="1E293B")
    hdr_font  = Font(color="94A3B8", bold=True, size=9, name="맑은 고딕")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 열 헤더 A(1)~O(15)
    COL_HDR = {
        1:"구분", 2:"WEB", 3:"시작일", 4:"종료일",
        5:"총예산\n(120%상향)", 6:"총예산\n수식", 7:"일예산\n수식", 8:"포인트\n(10)",
        9:"검색어", 10:"미션내용", 11:"정답",
        12:"힌트URL", 13:"업체명", 14:"일유입\n목표", 15:"글자수\n수식"
    }
    for col, title in COL_HDR.items():
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill  = hdr_fill
        cell.font  = hdr_font
        cell.alignment = hdr_align

    # 열 너비
    col_widths = {1:6,2:6,3:12,4:12,5:10,6:10,7:10,8:8,
                  9:12,10:42,11:12,12:36,13:18,14:8,15:8}
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    def cell_fill(color): return PatternFill("solid", fgColor=color)
    def thin_border(top_thick=False):
        t = Side(border_style="medium" if top_thick else "thin",
                 color="03C75A" if top_thick else "253048")
        s = Side(border_style="thin", color="253048")
        return Border(left=s, right=s, top=t, bottom=s)
    def camp_sep_border():
        thick = Side(border_style="medium", color="03C75A")
        thin  = Side(border_style="thin",   color="253048")
        return Border(left=thin, right=thin, top=thick, bottom=thin)

    # 병합 열 (1-based): A=1,C=3,D=4,E=5,F=6,G=7,H=8,N=14,O=15
    MERGE_COLS = {1,3,4,5,6,7,8,14,15}

    row_num = 2  # 1행=헤더, 데이터는 2행부터 (new file기준)
    for ci, camp in enumerate(campaigns):
        rows_data = camp.get("rows", [])
        camp_start_row = row_num
        is_first_camp = (ci == 0)

        for ri, rd in enumerate(rows_data):
            for col_idx in range(15):  # 0~14 → 열 1~15
                col_excel = col_idx + 1
                val = rd[col_idx] if col_idx < len(rd) else ""
                if val == "__merge__": val = ""

                if col_excel in MERGE_COLS and ri > 0:
                    continue

                c = ws.cell(row=row_num, column=col_excel, value=val)
                use_sep = (ri == 0 and not is_first_camp)
                c.border    = camp_sep_border() if use_sep else thin_border()
                c.font      = Font(name="맑은 고딕", size=9)
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                if col_excel == 1:   # A: 구분
                    c.font  = Font(name="맑은 고딕", color="64748B", size=9)
                    c.fill  = cell_fill("111827")
                elif col_excel == 2: # B: WEB
                    c.font  = Font(name="맑은 고딕", color="F59E0B", bold=True, size=9)
                    c.fill  = cell_fill("0D1220")
                elif col_excel in (3,4): # C,D: 날짜
                    c.font  = Font(name="맑은 고딕", color="03C75A", bold=True, size=9)
                    c.fill  = cell_fill("0D1A0D")
                elif col_excel in (5,6,7,8): # E,F,G,H: 수식
                    c.fill  = cell_fill("0D1220")
                    c.font  = Font(name="맑은 고딕", color="334155", size=9)
                elif col_excel == 9:  # I: 검색어
                    c.font  = Font(name="맑은 고딕", color="3B82F6", bold=True, size=9)
                    c.fill  = cell_fill("0D1220")
                elif col_excel == 10: # J: 미션내용
                    c.font  = Font(name="맑은 고딕", size=8)
                    c.fill  = cell_fill("0F172A")
                    c.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
                elif col_excel == 11: # K: 정답
                    c.font  = Font(name="맑은 고딕", size=9)
                    c.fill  = cell_fill("0D1220")
                elif col_excel == 12: # L: 힌트URL
                    c.font  = Font(name="맑은 고딕", color="3B82F6", size=8)
                    c.fill  = cell_fill("0D1220")
                elif col_excel == 13: # M: 업체명
                    c.font  = Font(name="맑은 고딕", color="A78BFA", bold=True, size=9)
                    c.fill  = cell_fill("0D1220")
                elif col_excel == 14: # N: 일유입목표
                    c.font  = Font(name="맑은 고딕", color="03C75A", bold=True, size=9)
                    c.fill  = cell_fill("0D1A0D")
                elif col_excel == 15: # O: 글자수
                    c.fill  = cell_fill("0D1220")
                    c.font  = Font(name="맑은 고딕", color="334155", size=9)

            row_num += 1

        # 행 높이
        for r in range(camp_start_row, camp_start_row + 5):
            ws.row_dimensions[r].height = 50

        # 병합: A,C,D,E,F,G,H,N,O
        end_row = camp_start_row + 4
        for mc in MERGE_COLS:
            try:
                ws.merge_cells(start_row=camp_start_row, start_column=mc,
                               end_row=end_row, end_column=mc)
                ws.cell(camp_start_row, mc).alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True)
            except Exception:
                pass

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read(), 200, {
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "Content-Disposition": "attachment; filename=campaign.xlsx"
    }


def _load_workbook_safe(path_or_bytes):
    """조건부 서식 XML 오류를 우회하여 워크북 로드"""
    import zipfile, re as _re
    from io import BytesIO as _BytesIO
    import openpyxl as _opx

    if isinstance(path_or_bytes, str):
        with open(path_or_bytes, 'rb') as _f:
            raw = _f.read()
    elif isinstance(path_or_bytes, (bytes, bytearray)):
        raw = bytes(path_or_bytes)
    else:
        raw = path_or_bytes.read()

    buf = _BytesIO()
    with zipfile.ZipFile(_BytesIO(raw), 'r') as zin:
        with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename.startswith('xl/worksheets/') and item.filename.endswith('.xml'):
                    text = data.decode('utf-8')
                    text = _re.sub(r'<conditionalFormatting[^>]*>.*?</conditionalFormatting>',
                                   '', text, flags=_re.DOTALL)
                    text = _re.sub(r'<conditionalFormatting[^/]*/>', '', text)
                    data = text.encode('utf-8')
                zout.writestr(item, data)
    buf.seek(0)
    return _opx.load_workbook(buf)


# ════════════════════════════════════════════
# Excel 기존 파일 채우기 (v15 구조)
# ════════════════════════════════════════════
@app.route("/api/automation/excel-fill", methods=["POST"])
def api_excel_fill():
    """기존 XLSX(템플릿) 파일에 캠페인 데이터 삽입 — v15 구조
    v15 컬럼: A(1)~O(15), 데이터 시작행=3, 헤더 2행
    병합: A,C,D,E,F,G,H(포인트수식),N(일유입),O(글자수수식)
    수식 보존: E,F,G,H,O (col_offset 4,5,6,7,14)
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import column_index_from_string, get_column_letter
    from io import BytesIO
    import json as _json
    import os

    file             = request.files.get("file")
    use_default      = request.form.get("use_default", "0") == "1"
    payload          = request.form.get("payload", "{}")
    start_row        = int(request.form.get("start_row", 3))
    start_col_letter = request.form.get("start_col", "A").upper().strip()

    data      = _json.loads(payload)
    campaigns = data.get("campaigns", [])

    DEFAULT_TEMPLATE = os.path.join(
        os.path.dirname(__file__),
        "static", "templates", "mission_template.xlsx")

    if use_default and os.path.exists(DEFAULT_TEMPLATE):
        wb = _load_workbook_safe(DEFAULT_TEMPLATE)
        out_name = "[SKP-재흥광고기획] 미션광고 리스트.xlsx"
    elif file:
        try:
            wb = _load_workbook_safe(BytesIO(file.read()))
        except Exception:
            file.seek(0)
            wb = openpyxl.load_workbook(BytesIO(file.read()), read_only=False, data_only=True)
        out_name = f"filled_{file.filename}"
    else:
        wb = openpyxl.Workbook()
        out_name = "filled_campaign.xlsx"

    ws = wb.active

    start_col_idx = column_index_from_string(start_col_letter)  # A=1

    # 병합 열 오프셋 (0-based): A=0,C=2,D=3,E=4,F=5,G=6,H=7,N=13,O=14
    MERGE_OFFSETS = [0, 2, 3, 4, 5, 6, 7, 13, 14]

    # 수식 보존 오프셋 (덮어쓰지 않음): E=4,F=5,G=6,H=7,O=14
    FORMULA_OFFSETS = {4, 5, 6, 7, 14}

    def thin_side(): return Side(border_style="thin", color="C8C8C8")
    def camp_border():
        thick = Side(border_style="medium", color="5EA15E")
        thin  = Side(border_style="thin",   color="C8C8C8")
        return Border(left=thin, right=thin, top=thick, bottom=thin)
    def normal_border():
        s = thin_side()
        return Border(left=s, right=s, top=s, bottom=s)

    row_num = start_row
    for ci, camp in enumerate(campaigns):
        rows_data = camp.get("rows", [])
        camp_start_row = row_num

        # 기존 병합 해제
        end_row_pre = row_num + 4
        ranges_to_unmerge = []
        for mr in list(ws.merged_cells.ranges):
            if (mr.min_row <= end_row_pre and mr.max_row >= row_num and
                    mr.min_col >= start_col_idx and mr.max_col <= start_col_idx + 14):
                ranges_to_unmerge.append(str(mr))
        for ref in ranges_to_unmerge:
            try: ws.unmerge_cells(ref)
            except: pass

        for ri, rd in enumerate(rows_data):
            for col_offset in range(15):  # 0~14
                col_excel = start_col_idx + col_offset
                val = rd[col_offset] if col_offset < len(rd) else ""
                if val == "__merge__": val = ""

                is_merge_col   = col_offset in MERGE_OFFSETS
                is_formula_col = col_offset in FORMULA_OFFSETS

                # 병합 열은 첫 행에만 쓰기
                if is_merge_col and ri > 0:
                    continue

                cell = ws.cell(row=row_num, column=col_excel)

                # 수식 열은 기존 값 보존 (덮어쓰지 않음)
                if is_formula_col:
                    pass  # 건드리지 않음
                else:
                    if val != "":
                        cell.value = val
                    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
                    if col_offset == 9:  # J: 미션내용 — 좌측 정렬
                        cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

                if ri == 0 and ci > 0:
                    cell.border = camp_border()
                elif ri == 0 and ci == 0:
                    cell.border = normal_border()

            row_num += 1

        # 행 높이 (v15 원본 파일 기준 낮게)
        for r in range(camp_start_row, camp_start_row + 5):
            ws.row_dimensions[r].height = 50

        # 병합 적용
        end_row = camp_start_row + 4
        for off in MERGE_OFFSETS:
            mc = start_col_idx + off
            try:
                ws.merge_cells(start_row=camp_start_row, start_column=mc,
                               end_row=end_row, end_column=mc)
                ws.cell(camp_start_row, mc).alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True)
            except Exception:
                pass

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read(), 200, {
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "Content-Disposition": f"attachment; filename={out_name}"
    }


# 하위 호환 라우트
@app.route("/products/add", methods=["POST"])
def add_product_compat():
    return redirect(url_for("index"))

@app.route("/keywords/add", methods=["POST"])
def add_keyword_compat():
    return redirect(url_for("index"))


with app.app_context():
    init_db()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(debug=False, host="0.0.0.0", port=port)
